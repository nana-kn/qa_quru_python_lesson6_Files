from zipfile import ZipFile
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook


with ZipFile('./resources/file_archive.zip', 'w') as zip_file:
    zip_file.write('./resources/Flats.pdf', 'Flats.pdf')
    zip_file.write('./resources/template_checklist.csv', 'template_checklist.csv')
    zip_file.write('./resources/template_checklist.xlsx', 'template_checklist.xlsx')


def test_csv_file():
    with ZipFile('./resources/file_archive.zip'):
        csv_file = csv.reader('template_checklist.csv')
        row_count = sum(1 for row in csv_file)
        assert row_count == 22, 'Количество строк не соответствует ожидаемому результату!'


def test_pdf_file():
    with ZipFile('./resources/file_archive.zip', 'r') as file_zip:
        pdf_file = file_zip.open('Flats.pdf')
        pdf_reader = PdfReader(pdf_file)
        count_pages = len(pdf_reader.pages)
        assert count_pages == 24, 'Количество страниц не соответствует ожидаемому результату!'


def test_xlsx_file():
    with ZipFile('./resources/file_archive.zip', 'r') as xlsx_file:
        xlsx_workbook = xlsx_file.open('template_checklist.xlsx')
        workbook = load_workbook(xlsx_workbook)
        page = workbook.sheetnames[1]
        assert page == 'Инструкция', "Вторая страница таблицы не соответствует ожидаемому результату!"




